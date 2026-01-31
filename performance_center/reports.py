# ===================================================================
# 📄 الملف: performance_center/reports.py
# 🧭 نظام التقارير الاحترافي — Performance Center Reports Engine V1.0
# 🚀 يدعم: PDF + Excel | Arabic RTL | Tajawal Font | Glass Design
# ===================================================================

from django.http import HttpResponse
from django.utils import timezone

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .models import PerformanceReview, ReviewAnswer


# ================================================================
# 🅰️ 1) إعداد خطوط PDF (Tajawal)
# ================================================================
# ملاحظة: تأكد من وضع الخط داخل static/fonts
try:
    pdfmetrics.registerFont(TTFont("Tajawal", "static/fonts/Tajawal-Regular.ttf"))
except:
    # fallback لو لم يتم تحميل الخط
    pass


# ================================================================
# 📝 2) دالة المساعدة — رسم نص عربي RTL
# ================================================================
def draw_rtl_text(c, text, x, y, size=12, bold=False):
    """
    ✨ دالة ذكية لعرض النصوص العربية RTL داخل PDF
    """
    c.setFont("Tajawal", size)
    c.drawRightString(x, y, text)


# ===================================================================
# 📘 3) generate_review_pdf — تقرير تقييم واحد كامل PDF
# ===================================================================
def generate_review_pdf(review_id):
    """
    📝 إنشاء تقرير PDF لتقييم أداء واحد (Self + Manager + HR)
    """
    review = PerformanceReview.objects.get(id=review_id)
    answers = ReviewAnswer.objects.filter(review=review).select_related("item")

    # إنشاء ملف PDF
    response = HttpResponse(content_type="application/pdf")
    filename = f"performance_review_{review.employee_id}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    c = canvas.Canvas(response, pagesize=A4)

    # هوامش
    margin_x = 190 * mm
    current_y = 270 * mm

    # ---------------------------------------------------------
    # 🧩 عنوان التقرير
    # ---------------------------------------------------------
    draw_rtl_text(c, "تقرير تقييم الأداء", margin_x, current_y, size=20)
    current_y -= 20

    draw_rtl_text(c, f"الموظف: {review.employee}", margin_x, current_y, size=13)
    current_y -= 10

    draw_rtl_text(c, f"القالب: {review.template.name}", margin_x, current_y, size=12)
    current_y -= 10

    draw_rtl_text(c, f"الفترة: {review.period_label}", margin_x, current_y, size=12)
    current_y -= 25

    # ---------------------------------------------------------
    # 🧾 جدول العناصر
    # ---------------------------------------------------------
    draw_rtl_text(c, "تفاصيل التقييم:", margin_x, current_y, size=15)
    current_y -= 15

    for ans in answers:
        if current_y < 40:
            c.showPage()
            current_y = 270 * mm

        draw_rtl_text(c, f"السؤال: {ans.item.question}", margin_x, current_y)
        current_y -= 8

        draw_rtl_text(
            c,
            f"درجة الموظف: {ans.self_score if ans.self_score else '—'}",
            margin_x,
            current_y,
        )
        current_y -= 8

        draw_rtl_text(
            c,
            f"درجة المدير: {ans.manager_score if ans.manager_score else '—'}",
            margin_x,
            current_y,
        )
        current_y -= 8

        draw_rtl_text(
            c,
            f"درجة HR: {ans.hr_score if ans.hr_score else '—'}",
            margin_x,
            current_y,
        )
        current_y -= 8

        draw_rtl_text(
            c,
            f"ملاحظات: {ans.hr_comment or ans.manager_comment or ans.self_comment or '—'}",
            margin_x,
            current_y,
        )
        current_y -= 15

    c.save()
    return response


# ===================================================================
# 📘 4) generate_employee_summary_pdf — تقرير موظف شامل
# ===================================================================
def generate_employee_summary_pdf(employee_id):
    """
    🧾 تقرير شامل لتقييمات موظف واحد (Multiple Reviews)
    """
    reviews = PerformanceReview.objects.filter(employee_id=employee_id)

    response = HttpResponse(content_type="application/pdf")
    filename = f"employee_summary_{employee_id}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    c = canvas.Canvas(response, pagesize=A4)

    margin_x = 190 * mm
    current_y = 270 * mm

    draw_rtl_text(c, f"تقرير شامل — الموظف رقم {employee_id}", margin_x, current_y, size=20)
    current_y -= 20

    for review in reviews:
        draw_rtl_text(c, f"- قالب: {review.template.name}", margin_x, current_y)
        current_y -= 10

        draw_rtl_text(c, f"  النتيجة النهائية: {review.final_score or '—'}", margin_x, current_y)
        current_y -= 10

        draw_rtl_text(c, f"  الحالة: {review.status}", margin_x, current_y)
        current_y -= 15

        if current_y < 40:
            c.showPage()
            current_y = 270 * mm

    c.save()
    return response


# ===================================================================
# 📘 5) export_reviews_excel — تصدير Excel لجميع التقييمات
# ===================================================================
def export_reviews_excel():
    """
    📊 إنشاء ملف Excel يحتوي جميع التقييمات في النظام
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Performance Reviews"

    headers = [
        "الموظف",
        "القالب",
        "الفترة",
        "الحالة",
        "النتيجة النهائية",
        "أخر تحديث"
    ]

    ws.append(headers)

    # تنسيق الرأس
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    # بيانات التقييم
    for review in PerformanceReview.objects.all():
        ws.append([
            str(review.employee),
            review.template.name,
            review.period_label,
            review.status,
            review.final_score,
            review.updated_at.strftime("%Y-%m-%d"),
        ])

    # تجهيز الاستجابة
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="performance_reviews.xlsx"'

    wb.save(response)
    return response
